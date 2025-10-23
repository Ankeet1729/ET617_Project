import openpyxl
import os
import re
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils.cell import coordinate_from_string # <-- This line is now fixed

def sanitize_filename(name):
    """Removes invalid characters from a string to make it a valid filename."""
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()

def get_video_name_for_row(sheet, row_num):
    """
    Finds the video name for a given row by looking for the last non-empty
    cell in column A at or before that row.
    """
    current_video_name = "Default_Video"
    for i in range(row_num, 0, -1):
        cell_value = sheet.cell(row=i, column=1).value
        if cell_value and str(cell_value).strip():
            current_video_name = str(cell_value)
            break
    return sanitize_filename(current_video_name)

def extract_images_by_video(xlsx_path):
    """
    Extracts all images from an XLSX file and sorts them into subdirectories
    based on their worksheet and a 'video name' derived from column A.
    """
    parent_folder = "extracted_content_by_video"
    if not os.path.exists(parent_folder):
        os.makedirs(parent_folder)
    print(f"Images will be saved in '{parent_folder}', sorted by sheet and video.")

    try:
        workbook = openpyxl.load_workbook(xlsx_path)
    except FileNotFoundError:
        print(f"Error: The file '{xlsx_path}' was not found.")
        return

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"\nProcessing sheet: '{sheet.title}'...")

        image_loader = SheetImageLoader(sheet)
        if not image_loader._images:
            print("  - No images found in this sheet.")
            continue

        video_image_counts = {}

        for cell_number in image_loader._images:
            try:
                row = coordinate_from_string(cell_number)[1]
                video_name = get_video_name_for_row(sheet, row)

                video_folder_path = os.path.join(parent_folder, sheet.title, video_name)
                if not os.path.exists(video_folder_path):
                    os.makedirs(video_folder_path)

                image = image_loader.get(cell_number)
                extension = image.format.lower() if image.format else 'png'

                current_count = video_image_counts.get(video_name, 0) + 1
                video_image_counts[video_name] = current_count
                new_filename = f"p{current_count}.{extension}"
                
                save_path = os.path.join(video_folder_path, new_filename)
                image.save(save_path)
                print(f"  - Found image at {cell_number}, saving to '{save_path}'")

            except Exception as e:
                print(f"  - Could not save image from cell {cell_number}. Reason: {e}")

# --- How to use the script ---
if __name__ == "__main__":
    excel_file = "extended_context.xlsx" 
    extract_images_by_video(excel_file)